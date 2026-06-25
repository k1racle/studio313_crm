"""
Генератор data-миграции для переноса данных из 'Расписание съемок.xlsx'.
Запускается один раз на машине разработчика:
    python backend/scripts/generate_schedule_migration.py
Результат: backend/apps/tasks/migrations/0006_import_schedule_data.py
"""

import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


def normalize_phone(value):
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        raw = str(int(value))
    else:
        raw = str(value)
    groups = re.findall(r'\d+', raw)
    for digits in groups:
        if len(digits) == 10 and digits.startswith('9'):
            digits = '7' + digits
        elif len(digits) == 11 and digits.startswith('8'):
            digits = '7' + digits[1:]
        if len(digits) == 11 and digits.startswith('7'):
            return digits
    all_digits = re.sub(r'\D', '', raw)
    return all_digits[:20]


def clean_email(value):
    if not value:
        return ''
    email = str(value).strip()
    if email.lower() in ('-', '—'):
        return ''
    if '@' not in email or ' ' in email:
        return ''
    return email


def clean_telegram(value):
    if not value:
        return ''
    tg = str(value).strip()
    if tg.lower() in ('-', '—'):
        return ''
    if ' ' in tg or '.' in tg or 'http' in tg.lower() or 'не ' in tg.lower():
        return ''
    if tg.startswith('@'):
        tg = tg[1:]
    return tg[:100]


def to_date(value):
    if isinstance(value, datetime):
        return value.date()
    return None


def to_datetime(value):
    if isinstance(value, datetime):
        return value
    return None


def join_notes(part1, part2):
    parts = [p for p in (part1, part2) if p]
    return '\n'.join(parts)


def cell_value(ws, row, col):
    if not col:
        return None
    value = ws.cell(row=row, column=col).value
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value in ('-', '—'):
            return None
    return value


def find_header_row(ws, key):
    key = key.strip().lower()
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            val = cell_value(ws, row, col)
            if val and str(val).strip().lower() == key:
                return row
    return None


def build_col_map(headers):
    return {str(h).strip(): idx + 1 for idx, h in enumerate(headers) if h}


def parse_clients(wb):
    ws = wb['База Yclients']
    header_idx = find_header_row(ws, 'ФИО')
    headers = [cell_value(ws, header_idx, col) for col in range(1, ws.max_column + 1)]
    col_map = build_col_map(headers)

    clients = []
    for row in range(header_idx + 1, ws.max_row + 1):
        name = cell_value(ws, row, col_map.get('ФИО'))
        if not name:
            continue
        phone = normalize_phone(cell_value(ws, row, col_map.get('Телефон')))
        email = clean_email(cell_value(ws, row, col_map.get('Email')))
        telegram = clean_telegram(cell_value(ws, row, col_map.get('Ник')))
        birthday = to_date(cell_value(ws, row, col_map.get('Дата рождения')))
        service = cell_value(ws, row, col_map.get('Услуга/Что покупали'))
        comment = cell_value(ws, row, col_map.get('Комментарий'))
        notes = join_notes(service, comment)
        clients.append({
            'name': name,
            'phone': phone,
            'email': email,
            'telegram': telegram,
            'birthday': birthday,
            'notes': notes,
        })
    return clients


def parse_tasks(wb):
    ws = wb['Очередь монтажей']
    header_idx = find_header_row(ws, 'ДАТА СЪЕМКИ')
    if not header_idx:
        return []
    headers = [cell_value(ws, header_idx, col) for col in range(1, ws.max_column + 1)]
    col_map = build_col_map(headers)

    status_map = {
        'готово': 'done',
        'правки': 'review',
        'отложено': 'canceled',
    }
    priority_map = {
        'важно': 'high',
    }

    tasks = []
    for row in range(header_idx + 1, ws.max_row + 1):
        title = cell_value(ws, row, col_map.get('ФИО клиента / название съемки'))
        if not title:
            continue
        shoot_date = to_datetime(cell_value(ws, row, col_map.get('ДАТА СЪЕМКИ')))
        fmt = cell_value(ws, row, col_map.get('Формат съемки'))
        deadline = to_datetime(cell_value(ws, row, col_map.get('ДЕДЛАЙН')))
        status_raw = cell_value(ws, row, col_map.get('СТАТУС'))
        priority_raw = cell_value(ws, row, col_map.get('ПРИОРИТЕТ'))
        comment = cell_value(ws, row, col_map.get('КОММЕНТАРИЙ'))
        editor_name = cell_value(ws, row, col_map.get('МОНТАЖЕР'))

        status = status_map.get((status_raw or '').strip().lower(), 'new')
        priority = priority_map.get((priority_raw or '').strip().lower(), 'medium')
        due_date = deadline or shoot_date

        tasks.append({
            'title': title,
            'description': join_notes(fmt, comment),
            'status': status,
            'priority': priority,
            'due_date': due_date,
            'client_name': title,
            'editor_name': editor_name,
        })
    return tasks


def format_value(value):
    if isinstance(value, str):
        return repr(value)
    if isinstance(value, date) and not isinstance(value, datetime):
        return f'datetime.date({value.year}, {value.month}, {value.day})'
    if isinstance(value, datetime):
        return f'make_aware(datetime.datetime({value.year}, {value.month}, {value.day}, {value.hour}, {value.minute}))'
    if value is None:
        return 'None'
    return repr(value)


def main():
    project_root = Path(__file__).resolve().parent.parent.parent
    excel_path = project_root / 'Расписание съемок.xlsx'
    wb = load_workbook(excel_path, data_only=True)

    clients = parse_clients(wb)
    tasks = parse_tasks(wb)

    lines = [
        '# Generated by backend/scripts/generate_schedule_migration.py',
        '# Data imported from "Расписание съемок.xlsx"',
        '',
        'import datetime',
        'from django.db import migrations',
        'from django.utils.timezone import make_aware',
        '',
        '',
        'CLIENTS = [',
    ]
    for client in clients:
        lines.append('    {')
        for key, value in client.items():
            lines.append(f'        {repr(key)}: {format_value(value)},')
        lines.append('    },')
    lines.append(']')
    lines.append('')
    lines.append('')
    lines.append('TASKS = [')
    for task in tasks:
        lines.append('    {')
        for key, value in task.items():
            lines.append(f'        {repr(key)}: {format_value(value)},')
        lines.append('    },')
    lines.append(']')
    lines.append('')
    lines.append('')
    lines.append('def import_schedule_data(apps, schema_editor):')
    lines.append('    Client = apps.get_model(\'clients\', \'Client\')')
    lines.append('    Task = apps.get_model(\'tasks\', \'Task\')')
    lines.append('    Project = apps.get_model(\'projects\', \'Project\')')
    lines.append('    User = apps.get_model(\'users\', \'User\')')
    lines.append('')
    lines.append('    for item in CLIENTS:')
    lines.append('        lookup = {}')
    lines.append('        if item.get(\'phone\'):')
    lines.append('            lookup[\'phone\'] = item[\'phone\']')
    lines.append('        if not lookup:')
    lines.append('            lookup[\'name__iexact\'] = item[\'name\']')
    lines.append('        Client.objects.get_or_create(defaults=item, **lookup)')
    lines.append('')
    lines.append('    project, _ = Project.objects.get_or_create(')
    lines.append('        name=\'Студия 313\',')
    lines.append('        defaults={\'description\': \'Клиенты и задачи Студии 313, импортированные из Excel\'},')
    lines.append('    )')
    lines.append('')
    lines.append('    def get_or_create_editor(name):')
    lines.append('        if not name:')
    lines.append('            return None')
    lines.append('        name = str(name).strip()')
    lines.append('        if name.lower() in (\'оутсорс\', \'аутсорс\', \'outsourcing\'):')
    lines.append('            return None')
    lines.append('        user = User.objects.filter(first_name__iexact=name).first()')
    lines.append('        if user:')
    lines.append('            return user')
    lines.append('        username = name.lower().replace(\' \', \'_\') + \'_mont\'')
    lines.append('        username = username[:150]')
    lines.append('        base_username = username')
    lines.append('        counter = 1')
    lines.append('        while User.objects.filter(username=username).exists():')
    lines.append('            suffix = f\'_{counter}\'')
    lines.append('            username = base_username[:150 - len(suffix)] + suffix')
    lines.append('            counter += 1')
    lines.append('        return User.objects.create(')
    lines.append('            username=username,')
    lines.append('            first_name=name,')
    lines.append('            role=\'staff\',')
    lines.append('            position=\'Монтажёр\',')
    lines.append('        )')
    lines.append('')
    lines.append('    for item in TASKS:')
    lines.append('        client = Client.objects.filter(name__iexact=item[\'client_name\']).first()')
    lines.append('        assignee = get_or_create_editor(item.pop(\'editor_name\', None))')
    lines.append('        item.pop(\'client_name\', None)')
    lines.append('        Task.objects.create(')
    lines.append('            project=project,')
    lines.append('            client=client,')
    lines.append('            assignee=assignee,')
    lines.append('            source=\'manual\',')
    lines.append('            **item')
    lines.append('        )')
    lines.append('')
    lines.append('')
    lines.append('def reverse_import(apps, schema_editor):')
    lines.append('    # Импортированные данные не удаляются автоматически, чтобы избежать потери связанных записей.')
    lines.append('    pass')
    lines.append('')
    lines.append('')
    lines.append('class Migration(migrations.Migration):')
    lines.append('    dependencies = [')
    lines.append('        (\'tasks\', \'0005_alter_task_source\'),')
    lines.append('        (\'clients\', \'0003_client_birthday\'),')
    lines.append('        (\'users\', \'0003_user_position_and_staff_role\'),')
    lines.append('    ]')
    lines.append('')
    lines.append('    operations = [')
    lines.append('        migrations.RunPython(import_schedule_data, reverse_import),')
    lines.append('    ]')
    lines.append('')

    migration_path = Path(__file__).resolve().parent.parent / 'apps' / 'tasks' / 'migrations' / '0006_import_schedule_data.py'
    migration_path.write_text('\n'.join(lines), encoding='utf-8')
    print(f'Миграция создана: {migration_path}')
    print(f'Клиентов: {len(clients)}, задач: {len(tasks)}')


if __name__ == '__main__':
    main()
